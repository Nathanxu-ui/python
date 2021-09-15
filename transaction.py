import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    print(wb)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)  # only cells
        corrected_price = cell.value * 0.95
        new_cell = sheet.cell(row, 4)
        new_cell.value = corrected_price

    values = Reference(sheet, min_row=2, max_row=sheet.max_row,
                       min_col=4, max_col=4)

    chart = BarChart()  # create a instance of barchart
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')  # create chart in the sheet, on e2 column

    wb.save(filename)


process_workbook("transactions.xlsx")

